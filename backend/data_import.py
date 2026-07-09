"""
知微能碳管理系统 — Excel 数据导入模块
支持从 Excel 文件导入系统数据，替换模拟数据/硬编码数据
"""
import os, json, sqlite3
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB = os.path.join(os.path.dirname(__file__), "energy_data.db")
IMPORT_DIR = os.path.join(os.path.dirname(__file__), "import_data")
os.makedirs(IMPORT_DIR, exist_ok=True)


def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


# ========== 导入模板定义 ==========
# 每个模板定义：表名、说明、字段列表(field_name, field_label, field_type, required, example)
TEMPLATES = {
    "device_groups": {
        "title": "耗能单元基础信息（生产单元→设备组→设备）",
        "desc": "定义系统中的耗能单元三级层级：生产单元（车间/产线）→ 设备组 → 单个设备。当前处于生产单元级。",
        "fields": [
            ("name", "单元名称", "文本", True, "A车间"),
            ("unit_type", "单元类型", "选项(production_unit/device_group/device)", True, "production_unit"),
            ("parent_name", "上级单元名称（最顶层留空）", "文本", False, ""),
            ("rated_kw", "额定功率(kW)", "数值", False, "500.0"),
            ("location", "位置", "文本", False, "1号厂房"),
            ("description", "描述", "文本", False, "空压车间"),
        ],
    },
    "production_data": {
        "title": "日产量数据",
        "desc": "各日的产品产量和产值数据，用于计算单位产品能耗和单位产值能耗",
        "fields": [
            ("date", "日期", "日期(YYYY-MM-DD)", True, "2026-06-01"),
            ("units_produced", "产量(件)", "整数", True, "120"),
            ("daily_output_value", "产值(万元)", "数值", True, "36.5"),
            ("source", "数据来源", "文本", False, "erp"),
        ],
    },
    "energy_records": {
        "title": "能耗时序数据",
        "desc": "全厂能耗时序记录，建议5分钟一条记录，可批量导入历史数据",
        "fields": [
            ("timestamp", "采集时间", "日期时间(YYYY-MM-DD HH:MM)", True, "2026-06-01 08:00"),
            ("total_energy_kwh", "累计电能(kWh)", "数值", True, "23456.7"),
            ("total_active_power_kw", "总有功功率(kW)", "数值", True, "320.5"),
            ("power_factor", "功率因数", "数值", False, "0.92"),
            ("frequency_hz", "频率(Hz)", "数值", False, "50.02"),
        ],
    },
    "device_energy": {
        "title": "设备能耗数据",
        "desc": "各设备/设备组的实时功率和能耗数据，可按时间批量导入",
        "fields": [
            ("timestamp", "采集时间", "日期时间(YYYY-MM-DD HH:MM)", True, "2026-06-01 08:00"),
            ("device_name", "设备名称", "文本", True, "双螺杆空压机组"),
            ("power_kw", "功率(kW)", "数值", True, "88.5"),
            ("energy_kwh", "时段电能(kWh)", "数值", True, "7.4"),
            ("current_a", "电流(A)", "数值", False, "135.2"),
            ("voltage_v", "电压(V)", "数值", False, "380.0"),
            ("power_factor", "功率因数", "数值", False, "0.91"),
        ],
    },
    "alarm_rules": {
        "title": "报警规则",
        "desc": "配置各设备组的监控报警规则，支持功率/电流/功率因数等指标",
        "fields": [
            ("device_group", "设备组", "文本", True, "双螺杆空压机组"),
            ("metric", "监控指标", "文本", True, "功率"),
            ("threshold", "阀值", "数值", True, "150"),
            ("operator", "比较符", "选项(gt/lt)", True, "gt"),
            ("level", "报警级别", "选项(warning/error/critical)", True, "warning"),
            ("message", "描述", "文本", False, "功率超限"),
            ("enabled", "启用", "整数(1/0)", False, "1"),
        ],
    },
    "suppliers": {
        "title": "供应商基础信息",
        "desc": "供应商清单管理，由系统管理员导入。碳评分和等级由系统根据提交数据自动计算。",
        "fields": [
            ("supplier_name", "供应商名称", "文本", True, "上海XX钢材有限公司"),
            ("supplier_code", "供应商编码", "文本", True, "SHTEEL001"),
            ("contact_person", "联系人", "文本", False, "张三"),
            ("contact_email", "联系邮箱", "文本", False, "zhang@steel.com"),
            ("region", "地区", "文本", False, "华东"),
        ],
    },
    "supplier_monthly_data": {
        "title": "供应商月度能碳数据",
        "desc": "供应商按月提交的原始能耗统计数据。系统自动计算碳排放量和碳评分，供应商只需填写各能源消耗量。",
        "fields": [
            ("supplier_code", "供应商编码", "文本", True, "SHTEEL001"),
            ("report_month", "报告月份", "日期(YYYY-MM)", True, "2026-06"),
            ("production_units", "产品产量(件)", "整数", True, "15000"),
            ("electricity_kwh", "外购电力(kWh)", "数值", True, "125000"),
            ("natural_gas_m3", "天然气(m³)", "数值", False, "5000"),
            ("steam_ton", "蒸汽(吨)", "数值", False, "200"),
            ("water_ton", "水(吨)", "数值", False, "3000"),
            ("diesel_l", "柴油(升)", "数值", False, "3000"),
            ("gasoline_l", "汽油(升)", "数值", False, "1000"),
            ("coal_ton", "煤炭(吨)", "数值", False, "0"),
            ("compressed_air_m3", "压缩空气(m³)", "数值", False, "0"),
            ("notes", "备注", "文本", False, "本月正常生产"),
        ],
    },
    "carbon_trade_prices": {
        "title": "碳交易价格趋势",
        "desc": "替换当前硬编码的碳价走势图数据，支持按日期导入价格",
        "fields": [
            ("date", "日期", "日期(YYYY-MM-DD)", True, "2026-06-01"),
            ("price", "价格(元/吨)", "数值", True, "58.5"),
            ("volume", "成交量(吨)", "数值", False, "35000"),
        ],
    },
    "ccer_projects": {
        "title": "CCER项目",
        "desc": "替换当前硬编码的CCER项目列表，支持导入碳减排项目",
        "fields": [
            ("name", "项目名称", "文本", True, "德耐尔光伏屋顶项目"),
            ("type", "项目类型", "选项(可再生能源/节能增效/碳汇)", True, "可再生能源"),
            ("status", "状态", "选项(已备案/开发中/已核证)", True, "已备案"),
            ("reduction", "预计减排量(吨)", "数值", True, "850"),
            ("progress", "进度(%)", "数值(0-100)", False, "75"),
        ],
    },
    "emission_factors": {
        "title": "碳排放因子",
        "desc": "配置各排放源的碳排放因子，用于碳排放核算（替换全局CF值）",
        "fields": [
            ("factor_name", "因子名称", "文本", True, "电力排放因子"),
            ("value", "因子值(kgCO₂/kWh)", "数值", True, "0.5566"),
            ("source", "来源/引用标准", "文本", False, "生态环境部2024年数据"),
            ("effective_year", "生效日期", "文本", False, "2026"),
        ],
    },
    "carbon_quotas": {
        "title": "碳配额",
        "desc": "各组织的年度碳配额和已使用量",
        "fields": [
            ("org_id", "组织ID", "整数", True, "1"),
            ("year", "年份", "整数(YYYY)", True, "2026"),
            ("total_quota_ton", "总配额(吨)", "数值", True, "15000"),
            ("used_ton", "已用(吨)", "数值", True, "8200"),
        ],
    },
    "optimization_tasks": {
        "title": "优化任务",
        "desc": "能效优化任务管理，导入后可在优化任务页面查看",
        "fields": [
            ("title", "任务名称", "文本", True, "空压机变频改造"),
            ("device_group", "设备组", "文本", True, "双螺杆空压机组"),
            ("expected_saving_kwh", "预估节能(kWh/年)", "数值", True, "15000"),
            ("status", "状态", "选项(待执行/执行中/已完成)", True, "待执行"),
        ],
    },
    "org_structure": {
        "title": "组织架构",
        "desc": "企业组织架构树，支持多级组织",
        "fields": [
            ("name", "组织名称", "文本", True, "德耐尔集团"),
            ("parent_name", "上级组织名称（根节点留空）", "文本", False, ""),
            ("level", "层级(0=根)", "整数", True, "0"),
        ],
    },
    "energy_flow_config": {
        "title": "能流分析配置",
        "desc": "替换当前硬编码的能流桑基图流向分配系数，可自定义各设备组的能耗流向比例",
        "fields": [
            ("from_node", "流入节点", "文本", True, "购入电力"),
            ("to_node", "流出节点", "文本", True, "空压机组"),
            ("ratio_pct", "分配比例(%)", "数值(0-100)", True, "28"),
        ],
    },
    "data_sources": {
        "title": "数据源配置",
        "desc": "管理系统中的所有数据采集来源，包括Excel、MQTT、HTTP、ERP等。导入后可在数据源管理页面查看配置",
        "fields": [
            ("source_name", "数据源名称", "文本", True, "1号电表"),
            ("source_type", "数据源类型", "选项(excel/mqtt/http/erp)", True, "excel"),
            ("config", "配置参数(JSON)", "文本", False, '{"note":"Excel人工上传"}'),
            ("status", "状态", "选项(active/inactive)", False, "active"),
            ("remark", "备注", "文本", False, "A车间总进线电表"),
        ],
    },
    "import_logs": {
        "title": "导入记录",
        "desc": "记录所有数据导入操作的历史，可手动补录导入记录用于审计",
        "fields": [
            ("table_name", "目标表名", "文本", True, "energy_records"),
            ("file_name", "文件名", "文本", True, "2026-06-07_能耗数据.xlsx"),
            ("records_imported", "导入记录数", "整数", True, "288"),
            ("status", "导入状态", "选项(success/partial/failed)", True, "success"),
            ("error_message", "错误信息", "文本", False, ""),
            ("imported_by", "导入人", "文本", False, "管理员"),
        ],
    },
}

TABLE_ORDER = [
    "device_groups", "org_structure", "production_data", "emission_factors", "carbon_quotas",
    "alarm_rules", "optimization_tasks",
    "suppliers", "supplier_monthly_data", "carbon_trade_prices", "ccer_projects", "energy_flow_config",
    "data_sources", "import_logs",
    "energy_records", "device_energy",
]


def get_schema(table_name):
    """获取指定表的字段定义"""
    return TEMPLATES.get(table_name, None)


def generate_template_excel(table_name):
    """生成指定表的空白导入模板 Excel"""
    schema = get_schema(table_name)
    if not schema:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = schema["title"]

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(schema["fields"]))
    title_cell = ws.cell(row=1, column=1, value=f"{schema['title']} — 导入模板")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="0d7377")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 说明行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(schema["fields"]))
    desc_cell = ws.cell(row=2, column=1, value=schema["desc"])
    desc_cell.font = Font(size=10, color="666666")
    desc_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 25

    # 表头行
    headers = [f"{f[1]}（{'必填' if f[3] else '选填'}）" for f in schema["fields"]]
    example_row = [f[4] for f in schema["fields"]]

    header_fill = PatternFill("solid", fgColor="e8eaed")
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 示例行
    for col, val in enumerate(example_row, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(size=10, color="999999", italic=True)
        cell.border = thin_border

    # 空数据行（从第5行开始）
    for col in range(1, len(schema["fields"]) + 1):
        cell = ws.cell(row=5, column=col, value="")
        cell.border = thin_border

    # 列宽
    for col in range(1, len(schema["fields"]) + 1):
        ws.column_dimensions[chr(64 + col)].width = max(len(headers[col - 1]) * 2, 18)

    return wb


def parse_and_validate(ws, table_name):
    """解析 Excel 工作表并验证数据"""
    schema = get_schema(table_name)
    if not schema:
        return None, "未知表名"

    fields = schema["fields"]
    start_row = 5  # 数据从第5行开始

    records = []
    errors = []

    for row_idx in range(start_row, ws.max_row + 1):
        record = {}
        row_empty = True
        for col_idx, (field_name, _, field_type, required, _) in enumerate(fields, 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None or str(cell_val).strip() == "":
                if required:
                    errors.append(f"第{row_idx}行「{field_name}」必填")
                record[field_name] = None
                continue
            row_empty = False
            val = str(cell_val).strip()
            # 类型转换
            if field_type == "数值":
                try:
                    record[field_name] = float(val)
                except ValueError:
                    errors.append(f"第{row_idx}行「{field_name}」应为数值，得到'{val}'")
            elif field_type == "整数":
                try:
                    record[field_name] = int(float(val))
                except ValueError:
                    errors.append(f"第{row_idx}行「{field_name}」应为整数，得到'{val}'")
            else:
                record[field_name] = val

        if row_empty:
            continue  # 跳过空行
        if not errors:
            records.append(record)

    return records, errors


def import_device_groups(records):
    """耗能单元导入（支持三级层级：生产单元→设备组→设备）"""
    conn = db(); c = conn.cursor()
    count = 0
    # 先清空重建（树形结构需要整体刷新）
    c.execute("DELETE FROM device_groups")
    name_cache = {}  # name → id
    # 先导入第一遍（建立所有节点）
    level_order = {"production_unit": 1, "device_group": 2, "device": 3}
    for r in sorted(records, key=lambda x: level_order.get(x.get("unit_type", "production_unit"), 1)):
        parent_id = None
        if r.get("parent_name") and r["parent_name"] in name_cache:
            parent_id = name_cache[r["parent_name"]]
        unit_type = r.get("unit_type", "production_unit")
        level = level_order.get(unit_type, 1)
        c.execute("INSERT INTO device_groups (name, type, rated_kw, location, count) VALUES (?,?,?,?,?)",
                  (r["name"], unit_type, r.get("rated_kw", 0) or 0,
                   r.get("location", ""), 1))
        new_id = c.lastrowid
        name_cache[r["name"]] = new_id
        count += 1
    conn.commit(); conn.close()
    return count


def import_production_data(records):
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        c.execute("INSERT OR REPLACE INTO production_data (date, units_produced, daily_output_value) VALUES (?,?,?)",
                  (r["date"], r["units_produced"], r["daily_output_value"]))
        count += c.rowcount
    conn.commit(); conn.close()
    return count


def import_energy_records(records):
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        c.execute("INSERT INTO energy_records (timestamp, total_energy_kwh, total_active_power_kw, power_factor, compressor_power_kw, milling_power_kw, cnc_power_kw, testbench_power_kw, aux_power_kw, phase_current_a, phase_voltage_v) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                  (r["timestamp"], r.get("total_energy_kwh"), r.get("total_active_power_kw"),
                   r.get("power_factor"), r.get("compressor_power_kw"), r.get("milling_power_kw"),
                   r.get("cnc_power_kw"), r.get("testbench_power_kw"), r.get("aux_power_kw"),
                   r.get("phase_current_a"), r.get("phase_voltage_v")))
        count += 1
    conn.commit(); conn.close()
    return count


def import_device_energy(records):
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        c.execute("INSERT INTO device_energy (timestamp, device_name, power_kw, energy_kwh, current_a, voltage_v, power_factor) VALUES (?,?,?,?,?,?,?)",
                  (r["timestamp"], r["device_name"], r["power_kw"], r["energy_kwh"],
                   r.get("current_a"), r.get("voltage_v"), r.get("power_factor")))
        count += 1
    conn.commit(); conn.close()
    return count


def import_alarm_rules(records):
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        c.execute("INSERT INTO alarm_rules (device_group, metric, threshold, operator, level, message, enabled) VALUES (?,?,?,?,?,?,?)",
                  (r["device_group"], r["metric"], r["threshold"], r.get("operator", "gt"),
                   r.get("level", "warning"), r.get("message", ""), r.get("enabled", 1)))
        count += 1
    conn.commit(); conn.close()
    return count


def import_suppliers(records):
    """供应商基础信息 — 插入suppliers表并自动创建账号"""
    from auth import hash_password
    conn = db(); c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS suppliers (id INTEGER PRIMARY KEY AUTOINCREMENT, supplier_code VARCHAR(50) UNIQUE, supplier_name VARCHAR(200), contact_person VARCHAR(100), contact_email VARCHAR(200), contact_phone VARCHAR(50), carbon_score REAL DEFAULT 0, carbon_level VARCHAR(10), annual_co2_ton REAL DEFAULT 0, status VARCHAR(20) DEFAULT 'active', source VARCHAR(20) DEFAULT 'manual', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS supplier_users (id INTEGER PRIMARY KEY AUTOINCREMENT, supplier_id INTEGER, username VARCHAR(100) UNIQUE, password_hash VARCHAR(128), api_key VARCHAR(64), must_change_password INTEGER DEFAULT 1, is_active INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    count = 0
    for r in records:
        c.execute("INSERT OR IGNORE INTO suppliers (supplier_code, supplier_name, contact_person, contact_email) VALUES (?,?,?,?)",
                  (r["supplier_code"], r.get("supplier_name"), r.get("contact_person"), r.get("contact_email")))
        if c.rowcount > 0:
            sid = c.lastrowid
            # 创建账号
            import random, string
            chars = string.ascii_uppercase + string.ascii_lowercase + string.digits
            pwd = "GP@" + ''.join(random.choice(string.ascii_uppercase) for _ in range(6)) + ''.join(random.choice(string.digits) for _ in range(2))
            c.execute("INSERT INTO supplier_users (supplier_id, username, password_hash) VALUES (?,?,?)",
                      (sid, f"sup_{sid}", hash_password(pwd)))
            count += 1
    conn.commit(); conn.close()
    return count


def import_carbon_trade_prices(records):
    """碳交易价格 — 插入carbon_trade_prices表"""
    conn = db(); c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS carbon_trade_prices (id INTEGER PRIMARY KEY AUTOINCREMENT, trade_date DATE, price REAL, volume INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    count = 0
    for r in records:
        c.execute("INSERT INTO carbon_trade_prices (trade_date, price, volume) VALUES (?,?,?)",
                  (r.get("date"), r.get("price"), r.get("volume")))
        count += 1
    conn.commit(); conn.close()
    return count


def import_ccer_projects(records):
    """CCER项目 — 插入ccer_projects表"""
    conn = db(); c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS ccer_projects (id INTEGER PRIMARY KEY AUTOINCREMENT, project_name VARCHAR(200), project_type VARCHAR(100), status VARCHAR(50), estimated_reduction REAL, progress INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    count = 0
    for r in records:
        c.execute("INSERT INTO ccer_projects (project_name, project_type, status, estimated_reduction, progress) VALUES (?,?,?,?,?)",
                  (r.get("name"), r.get("type"), r.get("status"), r.get("reduction"), r.get("progress")))
        count += 1
    conn.commit(); conn.close()
    return count
    return len(records)


def import_emission_factors(records):
    """排放因子 — 插入emission_factors表"""
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        try:
            c.execute("INSERT INTO emission_factors (factor_type, value, effective_year) VALUES (?,?,?)",
                      (r.get("factor_name") or r.get("factor_type"), r.get("value"), r.get("effective_year") or 2026))
            count += 1
        except Exception as e:
            print(f'  ⚠️ emission_factors插入失败: {e}')
    conn.commit(); conn.close()
    return count


def import_carbon_quotas(records):
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        c.execute("INSERT OR REPLACE INTO carbon_quotas (org_id, year, total_quota_ton, used_ton) VALUES (?,?,?,?)",
                  (r["org_id"], r["year"], r["total_quota_ton"], r["used_ton"]))
        count += 1
    conn.commit(); conn.close()
    return count


def import_optimization_tasks(records):
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        c.execute("INSERT INTO optimization_tasks (title, device_group, expected_saving_kwh, status) VALUES (?,?,?,?)",
                  (r["title"], r["device_group"], r["expected_saving_kwh"], r["status"]))
        count += 1
    conn.commit(); conn.close()
    return count


def import_supplier_monthly_data(records):
    """供应商月度能耗数据 — 只存原始统计值，不存评分"""
    conn = db(); c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS supplier_carbon_data (id INTEGER PRIMARY KEY AUTOINCREMENT, supplier_code VARCHAR(50), supplier_id INTEGER, report_month DATE, production_units INTEGER DEFAULT 0, electricity_kwh REAL DEFAULT 0, natural_gas_m3 REAL DEFAULT 0, steam_ton REAL DEFAULT 0, water_ton REAL DEFAULT 0, diesel_l REAL DEFAULT 0, gasoline_l REAL DEFAULT 0, coal_ton REAL DEFAULT 0, compressed_air_m3 REAL DEFAULT 0, calculated_co2_kg REAL DEFAULT 0, carbon_score REAL DEFAULT 0, carbon_level VARCHAR(10), notes TEXT, data_source VARCHAR(20) DEFAULT 'excel', submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, UNIQUE(supplier_code, report_month))")
    count = 0
    # 排放因子（系统统一配置）
    factors = {"electricity_kwh": 0.5566, "natural_gas_m3": 2.16, "steam_ton": 276.0,
               "diesel_l": 2.63, "gasoline_l": 2.30, "coal_ton": 2060.0, "compressed_air_m3": 0.0,
               "water_ton": 0.0}
    for r in records:
        # 计算碳排放（kg CO2）
        co2 = 0.0
        for field, factor in factors.items():
            val = r.get(field, 0) or 0
            co2 += float(val) * factor
        # 碳评分 = 基础分(100) - 排放强度系数
        units = r.get("production_units", 1) or 1
        intensity = co2 / float(units)  # kgCO2/件
        score = max(0, min(100, round(100 - intensity * 0.5, 1)))
        level = "A" if score >= 80 else ("B" if score >= 60 else ("C" if score >= 40 else "D"))
        # 查找对应的 supplier_id
        sid = None
        srow = c.execute("SELECT id FROM suppliers WHERE supplier_code=?", (r["supplier_code"],)).fetchone()
        if srow: sid = srow['id']
        c.execute("""INSERT OR REPLACE INTO supplier_carbon_data
            (supplier_code, supplier_id, report_month, production_units, electricity_kwh, natural_gas_m3,
             steam_ton, water_ton, diesel_l, gasoline_l, coal_ton, compressed_air_m3,
             calculated_co2_kg, carbon_score, carbon_level, notes, data_source)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (r["supplier_code"], sid, r["report_month"] + "-01", r.get("production_units", 0),
                   r.get("electricity_kwh", 0), r.get("natural_gas_m3", 0),
                   r.get("steam_ton", 0), r.get("water_ton", 0),
                   r.get("diesel_l", 0), r.get("gasoline_l", 0),
                   r.get("coal_ton", 0), r.get("compressed_air_m3", 0),
                   round(co2, 2), score, level, r.get("notes", ""), "excel"))
        count += 1
    conn.commit(); conn.close()
    return count


def import_org_structure(records):
    conn = db(); c = conn.cursor()
    count = 0
    # 先清空再导入（结构简单）
    c.execute("DELETE FROM org_structure")
    parent_cache = {}
    for r in records:
        parent_id = None
        if r.get("parent_name"):
            if r["parent_name"] in parent_cache:
                parent_id = parent_cache[r["parent_name"]]
        c.execute("INSERT INTO org_structure (name, parent_id, level) VALUES (?,?,?)",
                  (r["name"], parent_id, r["level"]))
        new_id = c.lastrowid
        parent_cache[r["name"]] = new_id
        count += 1
    conn.commit(); conn.close()
    return count


def import_energy_flow_config(records):
    """能流分析配置 — 存为JSON（替换硬编码流向比例）"""
    path = os.path.join(IMPORT_DIR, "energy_flow_config.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    return len(records)


def import_data_sources(records):
    """数据源配置 — 写入DB"""
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        config_json = r.get("config") or "{}"
        if isinstance(config_json, str):
            try: json.loads(config_json)
            except: config_json = "{}"
        c.execute("INSERT INTO data_sources (source_name, source_type, config, status) VALUES (?,?,?,?)",
                  (r["source_name"], r["source_type"], config_json, r.get("status", "active")))
        count += 1
    conn.commit(); conn.close()
    return count


def import_import_logs(records):
    """导入记录 — 写入DB"""
    conn = db(); c = conn.cursor()
    count = 0
    for r in records:
        c.execute("INSERT INTO import_logs (table_name, file_name, records_imported, status, error_message, imported_by) VALUES (?,?,?,?,?,?)",
                  (r["table_name"], r["file_name"], r["records_imported"],
                   r["status"], r.get("error_message", ""), r.get("imported_by", "管理员")))
        count += 1
    conn.commit(); conn.close()
    return count


IMPORTERS = {
    "device_groups": import_device_groups,
    "production_data": import_production_data,
    "energy_records": import_energy_records,
    "device_energy": import_device_energy,
    "alarm_rules": import_alarm_rules,
    "suppliers": import_suppliers,
    "carbon_trade_prices": import_carbon_trade_prices,
    "ccer_projects": import_ccer_projects,
    "emission_factors": import_emission_factors,
    "carbon_quotas": import_carbon_quotas,
    "optimization_tasks": import_optimization_tasks,
    "supplier_monthly_data": import_supplier_monthly_data,
    "org_structure": import_org_structure,
    "energy_flow_config": import_energy_flow_config,
    "data_sources": import_data_sources,
    "import_logs": import_import_logs,
}


def list_importable_tables():
    """列出所有可导入的表"""
    result = []
    for key in TABLE_ORDER:
        t = TEMPLATES[key]
        result.append({
            "table_name": key,
            "title": t["title"],
            "desc": t["desc"],
            "field_count": len(t["fields"]),
        })
    return result


def import_excel(file_path, table_name):
    """导入 Excel 文件到指定表"""
    schema = get_schema(table_name)
    if not schema:
        return None, f"不支持的表名：{table_name}，支持的：{list(TEMPLATES.keys())}"

    try:
        wb = load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        return None, f"Excel 文件解析失败：{str(e)}"

    records, errors = parse_and_validate(ws, table_name)
    if errors:
        return None, errors

    if not records:
        return None, "Excel 中没有有效数据行"

    importer = IMPORTERS.get(table_name)
    if not importer:
        return None, f"未找到 {table_name} 的导入处理器"

    try:
        count = importer(records)
        return {"imported": count, "table": table_name, "title": schema["title"]}, None
    except Exception as e:
        return None, f"导入 {table_name} 失败：{str(e)}"
